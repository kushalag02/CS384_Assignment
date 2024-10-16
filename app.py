import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill

def split_roll_name(input_file, output_file):
    # Split the input file
    df = pd.read_csv(input_file)

    # Split the 'Roll' column into 'Roll Number' and 'Name'
    df[['Roll Number', 'Name']] = df['Roll'].str.split(' ', n=1, expand=True)

    # Drop the old 'Roll' column
    df = df.drop(columns=['Roll'])

    df.to_csv(output_file, index=False)


def load_student_list(file): # Accessing the file
    students = {}
    with open(file, 'r') as f:
        for line in f:
            roll, name = line.strip().split(' ', 1)
            students[roll] = name
    return students

def load_dates(file): # Made changes according to the format of dates.txt
    with open(file, 'r') as f:
        dates = f.read().strip().split(', ')
    return dates

def process_attendance(file):
    # Process the attendance file
    data = pd.read_csv(file)
    data['Timestamp'] = pd.to_datetime(data['Timestamp'], format='%d-%m-%Y %H:%M')
    return data

def get_attendance_status(attendance_data, dates, students): # Processing attendence data by iterating
    attendance_summary = {roll: {date: 0 for date in dates} for roll in students.keys()}

    for roll in students.keys():
        for date in dates:
            date_formatted = datetime.strptime(date, '%d/%m/%Y').strftime('%Y-%m-%d')
            lecture_time_start = datetime.strptime(f"{date_formatted} 18:00", '%Y-%m-%d %H:%M')
            lecture_time_end = datetime.strptime(f"{date_formatted} 20:00", '%Y-%m-%d %H:%M')

            # Filter attendance for the student within the lecture time window
            student_attendance = attendance_data[
                (attendance_data['Roll Number'] == roll) & 
                (attendance_data['Timestamp'] >= lecture_time_start) & 
                (attendance_data['Timestamp'] <= lecture_time_end)
            ]

            # Assign attendance status
            if len(student_attendance) == 2:
                attendance_summary[roll][date] = 2  # Full attendance
            elif len(student_attendance) == 1:
                attendance_summary[roll][date] = 1  # Partial attendance
            elif len(student_attendance) > 2:
                attendance_summary[roll][date] = len(student_attendance)  # Mark as extra entries
            else:
                attendance_summary[roll][date] = 0  # Absent

    return attendance_summary

# Step 4: Generate Excel output with formatting
def generate_excel(attendance_summary, students, dates, output_file):
    # Create a DataFrame for the output with roll number and names as index
    df = pd.DataFrame(attendance_summary).T
    df.index = [f"{roll} {students[roll]}" for roll in df.index]  # Format index as 'Roll Number Name'
    df.columns = dates

    # Add the total attendance count from input_attendance.csv
    df['Total Attendance Marked'] = df.index.map(lambda x: total_attendance.get(x.split()[0], 0))

    # Add the sum of attendance values for each student (this replaces the valid attendance count)
    df['Sum of Attendance'] = df[dates].sum(axis=1)

    # Add the total number of classes taken (same for all students)
    df['Total Attendance Allowed'] = total_classes_taken

    # Add the 'Proxy' column (modulus of Total Attendance Count - Sum of Attendance)
    df['Proxy'] = (df['Total Attendance Marked'] - df['Sum of Attendance']).abs()

    # Adding extra columns for total attendance and proxy
    df['Total Attendance Marked'] = df.index.map(lambda x: total_attendance.get(x.split()[0], 0))
    df['Sum of Attendance'] = df[dates].sum(axis=1)
    df['Total Attendance Allowed'] = total_classes_taken
    df['Proxy'] = (df['Total Attendance Marked'] - df['Sum of Attendance']).abs()

    # Write to Excel
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df.to_excel(writer, sheet_name='Attendance', index=True)

    # Apply formatting
    worksheet = writer.sheets['Attendance']

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for row in worksheet.iter_rows(min_row=2, min_col=2, max_row=len(df)+1, max_col=len(dates)+1):
        for cell in row:
            if cell.value > 2:
                cell.fill = red_fill
            elif cell.value == 1:
                cell.fill = yellow_fill
            elif cell.value == 2:
                cell.fill = green_fill

    writer.close()

def main():
    split_roll_name('input_attendance.csv','input_attendance_processed.csv')
    students = load_student_list('stud_list.txt')  # Load roll numbers and names
    dates = load_dates('dates.txt')  # Modified to handle comma-separated dates
    attendance_data = process_attendance('input_attendance_processed.csv')

    # Get attendance summary
    attendance_summary = get_attendance_status(attendance_data, dates, students)
    generate_excel(attendance_summary, students, dates, 'output_excel.xlsx')

if __name__ == "__main__":
    main()
